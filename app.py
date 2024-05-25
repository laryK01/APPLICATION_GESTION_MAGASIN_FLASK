from io import BytesIO
import openpyxl
import plotly.express as px
import plotly.graph_objs as go
from flask import Flask, render_template, request, url_for, redirect, make_response, flash
from flask_mysqldb import MySQL
from flask_paginate import Pagination, get_page_parameter
from fpdf import FPDF

app = Flask(__name__)
app.secret_key = "flash message"

app.config['MYSQL_HOST'] ="localhost"
app.config['MYSQL_USER'] ="root"
app.config['MYSQL_PASSWORD'] =""
app.config['MYSQL_DB'] ="azza"

mysql = MySQL(app)


#@app.route('/data')
def get_data(year):
    cur = mysql.connection.cursor()
    cur.execute("""SELECT YEAR(Datevente) as annee, MONTH(Datevente) as mois, COUNT(*) as vente_total
                FROM vente WHERE YEAR(Datevente) = %s
                GROUP BY YEAR(Datevente), MONTH(Datevente)
                ORDER BY  annee, mois  """, (year,))
    data = cur.fetchall()
    data_vente = []
    for item in data:
        data_vente.append(item[2])

    cur.close()
    return (data_vente)


# Fonction pour obtenir les données de vente totales pour chaque année
def get_ventes_totales_par_annee():
    cur = mysql.connection.cursor()
    cur.execute("""SELECT YEAR(Datevente) as annee, COUNT(*) as vente_totale
                   FROM vente
                   GROUP BY YEAR(Datevente)
                   ORDER BY annee""")
    data = cur.fetchall()
    cur.close()
    return data


# fonction pour les ventes par catégories
def get_ventes_par_categorie():
    cur = mysql.connection.cursor()
    cur.execute("""SELECT CatProduit, COUNT(*) as vente_totale
                   FROM produit
                   GROUP BY CatProduit""")
    data = cur.fetchall()
    cur.close()
    return data

#meilleur produit


def get_produit():
    cur = mysql.connection.cursor()
    cur.execute("""SELECT NomProduit, COUNT(NomProduit)
                    FROM produit
                    GROUP BY NomProduit
                    ORDER BY NomProduit DESC

                    """)
    data = cur.fetchall()
    cur.close()
    return data

#10 meilleurs clients
def get_client():
    cur = mysql.connection.cursor()
    cur.execute("""SELECT NomClient, COUNT(*)               
                FROM vente
                GROUP BY NomClient
                ORDER BY Prixtotal DESC""")
    data = cur.fetchall()
    cur.close()
    return data




@app.route('/')
def index():
    cur = mysql.connection.cursor()
    #today = date.today()
    cur.execute('SELECT  COUNT(idProduit) from produit ')
    total_Produit = cur.fetchone()[0]



    # Récupérer total_sale
    cur.execute('SELECT COUNT(IdVente) FROM vente')
    total_sale = cur.fetchone()[0]
    
    if total_sale is None:
        total_sale = 0


    # Récupérer le revenu du jour
    #today = date.today()
    cur.execute('SELECT Count(NomClient) FROM vente ')
    total_client = cur.fetchone()[0]

    #if today_revenue is None:
        #today_revenue = 0

    
    # Récupérer le revenu total
    cur.execute('SELECT SUM(Prixtotal) FROM vente')
    total_revenue = cur.fetchone()[0]

    if total_revenue is None:
        total_revenue = 0

    cur.execute("""SELECT YEAR(Datevente) as annee, MONTH(Datevente) as mois, COUNT(*) as vente_total
                FROM vente
                GROUP BY YEAR(Datevente), MONTH(Datevente)
                ORDER BY  annee, mois  """)
    
    total_vente = cur.fetchall()
    #print(total_vente)

    vente_2022 = get_data(2022)
    vente_2023= get_data(2023)
    vente_2024 = get_data(2024)

    Mois = ['Janvier', 'Fevrier', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre',
            'Decembre']
    fig_2022 = px.line(
        x=dict(title=' Mois'),
        y=dict(title=' vente_2022 '),
        labels=Mois,
        title='Comparaison des ventes par année',
        template='plotly_dark',
        color_discrete_sequence=px.colors.qualitative.Plotly,
    )


    fig_2022.update_layout(
        height=300,   # Ajuster la hauteur du diagramme
        width = 450,
    )


    fig_2022.add_trace(go.Scatter(x=Mois, y=vente_2022, name='vente 2022'))
    fig_2022.add_trace(go.Scatter(x=Mois, y=vente_2023, name='vente 2023'))
    fig_2022.add_trace(go.Scatter(x=Mois, y=vente_2024, name='vente 2024'))
    graph = fig_2022.to_html()
    bar = go.Figure()



    # Ajout courbe d'évolution

    # Récupérer les données de vente totales pour chaque année
    ventes_par_annee = get_ventes_totales_par_annee()

    # Extraire les années et les ventes totales
    annees = [vente[0] for vente in ventes_par_annee]
    ventes_totales = [vente[1] for vente in ventes_par_annee]

    # Créer la trace de ligne pour les ventes totales par année
    trace_ventes_totales = go.Scatter(x=annees, y=ventes_totales, mode='lines+markers', name='Ventes totales')


    # Récupérer les données de vente totales pour chaque année
    ventes_par_annee = get_ventes_totales_par_annee()

    # Extraire les années et les ventes totales
    annees = [vente[0] for vente in ventes_par_annee]
    ventes_totales = [vente[1] for vente in ventes_par_annee]

    # Créer la trace de ligne pour les ventes totales par année
    trace_ventes_totales = go.Scatter(x=annees, y=ventes_totales, mode='lines+markers', name='Ventes totales')

    # Mise en forme du graphique
    layout = go.Layout(
        title='Évolution des ventes totales par année',
        xaxis=dict(title='Année'),
        yaxis=dict(title='Ventes totales'),
        plot_bgcolor='black',  # Définir la couleur de fond en noir
        paper_bgcolor='black',  # Définir la couleur de fond du papier en noir
        font=dict(color='white'),  # Définir la couleur du texte en blanc

    )
    layout.update(
        height=300,  # Ajuster la hauteur du diagramme
        width=450,
    )

    # Créer la figure
    fig = go.Figure(data=[trace_ventes_totales], layout=layout)

    # Convertir la figure en code HTML
    graph_html = fig.to_html(include_plotlyjs=False, full_html=False)


    #diagramme en cascade

    # Récupérer les données de ventes par catégorie
    ventes_par_categorie = get_ventes_par_categorie()

    # Trier les données par ventes totales (du plus petit au plus grand)
    ventes_par_categorie = sorted(ventes_par_categorie, key=lambda x: x[1])

    # Extraire les catégories et les ventes totales
    categories = [vente[0] for vente in ventes_par_categorie]
    ventes_totales = [vente[1] for vente in ventes_par_categorie]

    # Créer la trace de diagramme en cascade
    trace_ventes_par_categorie = go.Waterfall(
        x=categories,
        y=ventes_totales,
        orientation='v',  # Orientation verticale
        decreasing={'marker': {'color': 'red'}},  # Couleur des valeurs décroissantes
        increasing={'marker': {'color': 'red'}},  # Couleur des valeurs croissantes
        totals={'marker': {'color': 'red'}},  # Couleur du total
        name='Ventes par catégorie'
    )
    layout.update(
        height=400,  # Ajuster la hauteur du diagramme
        width=450,
    )

    # Mise en forme du graphique
    layout = go.Layout(
        title='Répartition des ventes par catégorie',
        xaxis=dict(title='Catégorie'),
        yaxis=dict(title='Ventes totales'),
        plot_bgcolor='black',  # Définir la couleur de fond en noir
        paper_bgcolor='black',  # Définir la couleur de fond du papier en noir
        font=dict(color='white'),  # Définir la couleur du texte en blanc
    )

    # Créer la figure
    fig = go.Figure(data=[trace_ventes_par_categorie], layout=layout)

    # Convertir la figure en code HTML
    cat_html = fig.to_html(include_plotlyjs=False, full_html=False)




    #diagramme circulaire

    # Récupérer les données sur les ventes par produit
    data_produits = get_produit()

    # Trier les données par nombre de ventes (du plus grand au plus petit) et limiter aux 10 premiers produits
    data_produits = sorted(data_produits, key=lambda x: x[1], reverse=True)[:10]

    # Extraire les noms des produits et le nombre de ventes pour chaque produit
    noms_produits = [produit[0] for produit in data_produits]
    ventes_produits = [produit[1] for produit in data_produits]

    # Créer la trace de diagramme circulaire
    trace_ventes_produits = go.Pie(
        labels=noms_produits,
        values=ventes_produits,
        hole=0.5,  # Taille du trou central
        name='Ventes par Produit'
    )
    layout.update(
        height=400,  # Ajuster la hauteur du diagramme
        width=350,
    )

    # Mise en forme du graphique
    layout = go.Layout(
        title='10 meilleures produits',
        plot_bgcolor='black',  # Couleur de fond du graphique
        paper_bgcolor='black',  # Couleur de fond du papier
        font=dict(color='white'),  # Couleur du texte
    )

    # Créer la figure
    fig = go.Figure(data=[trace_ventes_produits], layout=layout)

    prod_html = fig.to_html(include_plotlyjs=False, full_html=False)


    #meilleur client

    # Récupérer les données des clients
    data_clients = get_client()

    # Trier les données par le nombre d'achats (du plus grand au plus petit) et limiter aux 10 premiers clients
    data_clients = sorted(data_clients, key=lambda x: x[1], reverse=True)[:10]

    # Extraire les noms des clients et le nombre d'achats pour chaque client
    noms_clients = [client[0] for client in data_clients]
    achats_clients = [client[1] for client in data_clients]

    # Créer la trace de diagramme à barres
    trace_ventes_clients = go.Bar(
        x=noms_clients,
        y=achats_clients,
        name='Nombre d\'achats'
    )

    # Mise en forme du graphique
    layout = go.Layout(
        title='10 meilleurs clients',
        xaxis=dict(title='Clients'),
        yaxis=dict(title='Nombre d\'achats'),
        plot_bgcolor='black',  # Couleur de fond du graphique
        paper_bgcolor='black',  # Couleur de fond du papier
        font=dict(color='white'),  # Couleur du texte
    )

    # Créer la figure
    fig = go.Figure(data=[trace_ventes_clients], layout=layout)

    # Convertir la figure en code HTML
    clients_html = fig.to_html(include_plotlyjs=False, full_html=False)


    return render_template('index.html', total_Produit = total_Produit, total_sale=total_sale, total_client = total_client, total_revenue = total_revenue, graph = graph, graphe=graph_html, cat=cat_html, produit = prod_html, client = clients_html )



######### ADD ##########

@app.route('/add_from_produit')
def add_from_produit():
    return render_template('add_from_produit.html')

@app.route('/add_from_magasin')
def add_from_magasin():
    return render_template('add_from_magasin.html')

@app.route('/add_from_stock')
def add_from_stock():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produit")
    produit = cur.fetchall()
    cur.execute("SELECT * FROM magasin")
    magasin = cur.fetchall()
    cur.close()
    return render_template('add_from_stock.html', magasin=magasin, produit=produit)

#############################################################################################

#### produit ####

@app.route('/produit')
def list_produit():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 5  # Nombre d'éléments par page
    offset = (page - 1) * per_page

    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM produit")
    total_count = cur.fetchone()[0]
    cur.execute("SELECT * FROM produit LIMIT %s OFFSET %s", (per_page, offset))
    data = cur.fetchall()
    cur.close()

    pagination = Pagination(page=page, total=total_count, per_page=per_page, css_framework='bootstrap5')

    return render_template('list_produit.html', produit=data, pagination=pagination)

@app.route('/add_produit', methods=["POST"])
def add_produit():
    # Obtenir les données du formulaire
    nom = request.form['nom']
    categorie = request.form['categorie']
    prix = request.form['prix']

    # Établir une connexion à la base de données
    cur = mysql.connection.cursor()

    # Insérer les données dans la table 'produit'
    cur.execute("INSERT INTO produit (NomProduit, CatProduit, PrixUnitaire) VALUES (%s, %s, %s)", (nom, categorie, prix))

    # Valider la transaction
    mysql.connection.commit()

    # Fermer la connexion à la base de données
    cur.close()

    # Rediriger vers la page 'produit'
    return redirect(url_for('list_produit'))

@app.route('/upd_produit/<int:id>', methods=["GET", "POST"])
def upd_produit(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produit WHERE IdProduit=%s", (id,))
    produit = cur.fetchone()
    cur.close()
    
    if request.method == "POST":
        nom = request.form['nom']
        categorie= request.form['categorie']
        prix = request.form['prix']
        
        cur = mysql.connection.cursor()
        cur.execute("UPDATE produit SET NomProduit=%s, CatProduit=%s, PrixUnitaire=%s WHERE IdProduit=%s", (nom, categorie, prix, id))
        mysql.connection.commit()
        cur.close()
        
        return redirect(url_for('list_produit'))
    
    print("Produit : ", produit)
    if produit:
        return render_template('upd_produit.html', produit=produit)
    else:
        return "Produit introuvable"


@app.route('/delete_produit/<string:id_data>', methods=["GET"])
def delete_produit(id_data):
    cur = mysql.connection.cursor()
    cur.execute("SELECT IdProduit FROM produit WHERE IdProduit=%s", [id_data])
    stock_data = cur.fetchone()
    if stock_data:
        # si l'id existe dans la table stock, ne supprime pas la ligne correspondante dans la table produit
        flash("Impossible de supprimer le produit car il est présent dans le stock")
        return redirect(url_for('list_produit'))
    else:
        cur.execute("DELETE FROM produit WHERE IdProduit=%s", [id_data])
        mysql.connection.commit()
        cur.close()
        flash("Le produit a été supprimé avec succès")
        return redirect(url_for('list_produit'))


@app.route('/export_produit_pdf')
def export_produit_pdf():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produit")
    data = cur.fetchall()
    cur.close()

    # création du rapport en PDF
    pdf = FPDF()
    pdf.add_page()

    # ajout de l'en-tête
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des produit', 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 10, 'Ce document contient la liste de tous les produits', 0, 1, 'C')

    # ajout des titres des colonnes
    pdf.set_font('Arial', 'B', 16)
    col_width = pdf.w / 5.5
    row_height = pdf.font_size * 2
    for header in cur.description:
        pdf.cell(col_width, row_height, str(header[0]), border=1)

    # ajout des données de la base de données
    pdf.set_font('Arial', '', 10)
    for row in data:
        pdf.ln()
        for item in row:
            pdf.cell(col_width, row_height, str(item), border=1)

    # conversion du rapport en PDF en octets
    output = BytesIO()
    pdf.output(output)
    pdf_bytes = output.getvalue()

    # envoi du rapport en PDF au navigateur
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=magasins.pdf'
    return response

@app.route('/export_produit_excel')
def export_produit_excel():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produit")
    data = cur.fetchall()
    cur.close()

    # création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # ajout des titres des colonnes
    headers = [header[0] for header in cur.description]
    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])

    # ajout des données de la base de données
    for row in range(len(data)):
        for col in range(len(headers)):
            ws.cell(row=row+2, column=col+1, value=data[row][col])

    # conversion du classeur en octets
    output = BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    # envoi du classeur Excel au navigateur
    response = make_response(excel_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=produits.xlsx'
    return response


#### magasin ####

@app.route('/magasin')
def listemagasin():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 5
    offset = (page - 1) * per_page
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM magasin")
    total_count = cur.fetchone()[0]
    cur.execute("SELECT * FROM magasin LIMIT %s OFFSET %s", (per_page, offset))
    data = cur.fetchall()
    cur.close()
    pagination = Pagination(page=page, total=total_count, per_page=per_page, css_framework='bootstrap5')
    return render_template('list_magasin.html', magasin=data, pagination=pagination)

@app.route('/add_magasin', methods=["POST"])
def add_magasin():
    # Obtenir les données du formulaire
    nom = request.form['nom']
    adresse = request.form['adresse']
    Telephone = request.form['Telephone']
    mail = request.form['mail']


    # Établir une connexion à la base de données
    cur = mysql.connection.cursor()

    # Insérer les données dans la table 'magasin'
    cur.execute("INSERT INTO magasin (NomMagasin, AdresseMagasin, Telephone, mail) VALUES (%s, %s, %s, %s )", (nom, adresse, Telephone, mail ))

    # Valider la transaction
    mysql.connection.commit()

    # Fermer la connexion à la base de données
    cur.close()

    # Rediriger vers la page 'magasin'
    return redirect(url_for('listemagasin'))

@app.route('/upd_magasin/<int:id>', methods=["GET", "POST"])
def upd_magasin(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM magasin WHERE IdMagasin=%s", (id,))
    magasin = cur.fetchone()
    cur.close()
    
    if request.method == "POST":
        nom = request.form['nom']
        adresse = request.form['adresse']
        
        cur = mysql.connection.cursor()
        cur.execute("UPDATE magasin SET nom=%s, adresse=%s WHERE id=%s", (nom, adresse, id))
        mysql.connection.commit()
        cur.close()
        
        return redirect(url_for('listemagasin'))

    return render_template('upd_magasin.html', magasin=magasin)

@app.route('/delete_magasin/<string:id_data>', methods=["GET"])
def delete_magasin(id_data):
    cur = mysql.connection.cursor()
    cur.execute("SELECT IdMagasin FROM stock WHERE IdMagasin =%s", [id_data])
    stock_data = cur.fetchone()
    if stock_data:
        # si l'id existe dans la table stock, ne supprime pas la ligne correspondante dans la table magasin
        flash("Impossible de supprimer le magasin car il est présent dans le stock")
        return redirect(url_for('listemagasin'))
    else:
        cur.execute("DELETE FROM magasin WHERE IdMagasin=%s", [id_data])
        mysql.connection.commit()
        cur.close()
        flash("Le magasin a été supprimé avec succès")
        return redirect(url_for('listemagasin'))

@app.route('/export_magasin_pdf')
def export_magasin_pdf():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM magasin")
    data = cur.fetchall()
    cur.close()

    # création du rapport en PDF
    pdf = FPDF()
    pdf.add_page()

    # ajout de l'en-tête
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des magasins', 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 10, 'Ce document contient la liste de tous les magasins', 0, 1, 'C')

    # ajout des titres des colonnes
    pdf.set_font('Arial', 'B', 16)
    col_width = pdf.w / 5.5
    row_height = pdf.font_size * 2
    for header in cur.description:
        pdf.cell(col_width, row_height, str(header[0]), border=1)

    # ajout des données de la base de données
    pdf.set_font('Arial', '', 10)
    for row in data:
        pdf.ln()
        for item in row:
            pdf.cell(col_width, row_height, str(item), border=1)

    # conversion du rapport en PDF en octets
    output = BytesIO()
    pdf.output(output)
    pdf_bytes = output.getvalue()

    # envoi du rapport en PDF au navigateur
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=magasins.pdf'
    return response

@app.route('/export_magasin_excel')
def export_magasin_excel():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM magasin")
    data = cur.fetchall()
    cur.close()

    # création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # ajout des titres des colonnes
    headers = [header[0] for header in cur.description]
    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])

    # ajout des données de la base de données
    for row in range(len(data)):
        for col in range(len(headers)):
            ws.cell(row=row+2, column=col+1, value=data[row][col])

    # conversion du classeur en octets
    output = BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    # envoi du classeur Excel au navigateur
    response = make_response(excel_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=magasins.xlsx'
    return response

#### Vente ####
@app.route('/vente', methods=['GET', 'POST'])

def list_vente():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 5 # Nombre d'éléments par page
    offset = (page - 1) * per_page
    
    cur = mysql.connection.cursor()
    
    # Compter le nombre total d'éléments
    cur.execute("SELECT COUNT(*) FROM vente")
    total_count = cur.fetchone()[0]
    
    # Récupérer les données paginées
    cur.execute("SELECT V.*, P.NomProduit, P.CatProduit FROM vente V JOIN produit P ON P.IdProduit = V.IdProduit LIMIT %s OFFSET %s", (per_page, offset))
    data = cur.fetchall()
    
    cur.close()
    
    pagination = Pagination(page=page, total=total_count, per_page=per_page, css_framework='bootstrap5')
    
    return render_template('list_vente.html', vente=data, pagination=pagination)

#### stock ####

@app.route('/stock', methods=['GET', 'POST'])
def list_stock():
    # Récupération de la page actuelle depuis l'URL
    page = request.args.get(get_page_parameter(), type=int, default=1)

    # Nombre d'éléments à afficher par page
    per_page = 5

    # Définition des filtres
    id_produit = request.form.get('id_produit')
    # id_magasin = request.form.get('id_magasin')
    produit = get_nom_produit(id_produit) if id_produit else "Choose"
    # magasin = get_nom_magasin(id_magasin) if id_magasin else "Choose"

    # Construction de la requête SQL en fonction des filtres
    query = "SELECT COUNT(*) FROM produit JOIN stock ON produit.IdProduit = stock.IdProduit"
    condition = ""
    if id_produit :
        condition = " WHERE produit.IdProduit= %s "
        query += condition
        params = [id_produit]
        produit = get_nom_produit(id_produit)
        # magasin = get_nom_magasin(id_magasin)
    # elif id_produit:
    #     condition = " WHERE produit.IdProduit = %s"
    #     query += condition
    #     params = [id_produit]
    #     produit = get_nom_produit(id_produit)
    #     magasin = "Choose"
    
    else:
        params = []

    # Récupération du nombre total d'éléments dans la table 'stock' en fonction des filtres
    cur = mysql.connection.cursor()
    cur.execute(query, params)
    total_count = cur.fetchone()[0]

    # Calcul de l'offset à partir de la page et du nombre d'éléments par page
    offset = (page - 1) * per_page

    # Construction de la requête SQL pour récupérer les éléments à afficher en fonction des filtres
    query = "SELECT stock.IdStock, NomProduit, CatProduit, prixUnitaire FROM produit JOIN stock ON produit.IdProduit = stock.IdProduit "
    if condition:
        query += condition
    query += " LIMIT %s OFFSET %s"
    params = params + [per_page, offset]

    # Récupération des éléments à afficher pour la page actuelle en fonction des filtres
    cur.execute(query, params)
    stock_info = cur.fetchall()

    # Configuration de la pagination avec le nombre total d'éléments et le nombre d'éléments par page
    pagination = Pagination(page=page, total=total_count, per_page=per_page, css_framework='bootstrap5')

    cur.execute("SELECT * FROM produit")
    product_info = cur.fetchall()

    cur.execute("SELECT * FROM magasin")
    store_info = cur.fetchall()

    cur.close()

    # Rendu du template avec les données récupérées et la pagination
    return render_template('list_stock.html', products=product_info, stores=store_info, otto=stock_info, pagination=pagination, produit=produit)

@app.route('/add_stock', methods=["POST"])
def add_stock():

    # Obtenir les données du formulaire
    IdProduit = request.form['id_produit']
    IdMagasin= request.form['id_magasin']
    Quantitestock = request.form['quantite']

    # Établir une connexion à la base de données
    cur = mysql.connection.cursor()

    # Insérer les données dans la table 'stock'
    cur.execute("INSERT INTO stock (IdProduit, IdMagasin, Quantitestock) VALUES (%s, %s, %s)", (IdProduit, IdMagasin, Quantitestock))

    # Valider la transaction
    mysql.connection.commit()

    # Fermer la connexion à la base de données
    cur.close()

    # Rediriger vers la page 'magasin'
    return redirect(url_for('list_stock'))

@app.route('/upd_stock/<int:id>', methods=["GET", "POST"])
def upd_stock(id):
    cur = mysql.connection.cursor()
    cur.execute(
        "SELECT stock.Idstock, produit.NomProduit, produit.PrixUnitaire, magasin.NomMagasin, stock.Quantitestock FROM produit JOIN stock ON produit.IdProduit = stock.Idstock JOIN magasin ON magasin.IdMagasin = stock.IdMagasin WHERE stock.Idstock = %s",
        (id,))
    stock_info = cur.fetchone()
    cur.close()

    if request.method == "POST":
        Quantite = int(request.form['quantite'])
        cur = mysql.connection.cursor()
        cur.execute("SELECT Quantitestock FROM stock WHERE Idstock = %s", (id,))
        current_quantite = cur.fetchone()[0]
        new_quantite = current_quantite + Quantite
        cur.execute("UPDATE stock SET Quantitestock = %s WHERE Idstock = %s", (new_quantite, id))
        mysql.connection.commit()
        cur.close()
        return redirect(url_for('list_stock'))
    return render_template('upd_stock.html', stock_info=stock_info)


@app.route('/export_stock_pdf')
def export_stock_pdf():
    cur = mysql.connection.cursor()
    cur.execute("SELECT stock.id, produit.nom, produit.prix, magasin.nom, stock.quantite FROM produit JOIN stock ON produit.id = stock.id_produit JOIN magasin ON magasin.id = stock.id_magasin")
    data = cur.fetchall()
    cur.close()

    # création du rapport en PDF
    pdf = FPDF()
    pdf.add_page()

    # ajout de l'en-tête
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des stocks', 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 10, 'Ce document contient la liste de tous les stocks', 0, 1, 'C')

    # ajout des titres des colonnes
    pdf.set_font('Arial', 'B', 16)
    col_width = pdf.w / 5.5
    row_height = pdf.font_size * 2
    for header in cur.description:
        pdf.cell(col_width, row_height, str(header[0]), border=1)

    # ajout des données de la base de données
    pdf.set_font('Arial', '', 10)
    for row in data:
        pdf.ln()
        for item in row:
            pdf.cell(col_width, row_height, str(item), border=1)

    # conversion du rapport en PDF en octets
    output = BytesIO()
    pdf.output(output)
    pdf_bytes = output.getvalue()

    # envoi du rapport en PDF au navigateur
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=stocks.pdf'
    return response

@app.route('/export_stock_excel')
def export_stock_excel():
    cur = mysql.connection.cursor()
    cur.execute("SELECT stock.id, produit.nom, produit.prix, magasin.nom, stock.quantite FROM produit JOIN stock ON produit.id = stock.id_produit JOIN magasin ON magasin.id = stock.id_magasin")
    data = cur.fetchall()
    cur.close()

    # création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # ajout des titres des colonnes
    headers = [header[0] for header in cur.description]
    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])

    # ajout des données de la base de données
    for row in range(len(data)):
        for col in range(len(headers)):
            ws.cell(row=row+2, column=col+1, value=data[row][col])

    # conversion du classeur en octets
    output = BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    # envoi du classeur Excel au navigateur
    response = make_response(excel_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=stocks.xlsx'
    return response

#########################################################################

@app.route('/import_magasin_excel', methods=['POST'])
def import_magasin_excel():
    if 'fichier_excel' not in request.files:
        flash("Aucun fichier sélectionné")
        return redirect(url_for('listemagasin'))
    
    fichier = request.files['fichier_excel']
    
    # Lecture du fichier Excel
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active

    # Extraction des données
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Connexion à la base de données
    cur = mysql.connection.cursor()

    # Insertion des données dans la table SQL
    for row in data:
        query = "INSERT INTO magasin (nom, adresse) VALUES (%s, %s)"
        cur.execute(query, (row[1], row[2]))

    # Validation et enregistrement des modifications dans la base de données
    mysql.connection.commit()
    cur.close()
    flash("Importation réussie")
    return redirect(url_for('listemagasin'))



@app.route('/import_produit_excel', methods=['POST'])
def import_produit_excel():
    if 'fichier_excel' not in request.files:
        flash("Aucun fichier sélectionné")
        return redirect(url_for('list_produit'))

    fichier = request.files['fichier_excel']

    # Lecture du fichier Excel
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active

    # Extraction des données
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Connexion à la base de données
    cur = mysql.connection.cursor()

    # Insertion des données dans la table SQL
    for row in data:
        query = "INSERT INTO produit (nom, prix, description) VALUES (%s, %s, %s)"
        cur.execute(query, (row[1], row[2], row[3]))

    # Validation et enregistrement des modifications dans la base de données
    mysql.connection.commit()
    cur.close()

    flash("Importation réussie")
    return redirect(url_for('list_produit'))


# Correction de la fonction de récupération du nom de produit et de magasin
def get_nom_produit(id_produit):
    if id_produit:
        cur = mysql.connection.cursor()
        cur.execute("SELECT nom FROM produit WHERE id = %s", [id_produit])
        nom_produit = cur.fetchone()[0]
        cur.close()
        return nom_produit
    else:
        return ""

def get_nom_magasin(id_magasin):
    if id_magasin:
        cur = mysql.connection.cursor()
        cur.execute("SELECT nom FROM magasin WHERE id = %s", [id_magasin])
        nom_magasin = cur.fetchone()[0]
        cur.close()
        return nom_magasin
    else:
        return ""
    
# @app.route('/chart')
# def charts():
#     return render_template()



if __name__ == "__main__":
    app.run(debug=True, port=3000)